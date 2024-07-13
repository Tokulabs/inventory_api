from datetime import datetime
from decimal import Decimal

from django.db.models.functions.window import RowNumber
from django.utils import timezone
from django.db.models.functions.text import Upper

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from rest_framework.views import APIView
from sqlparse.sql import Case

from inventory_api.excel_manager import apply_styles_to_cells

from .serializers import (
    Inventory, Invoice, Customer
)
from rest_framework.response import Response
from inventory_api.custom_methods import IsAuthenticatedCustom
from inventory_api.utils import create_terminals_report, create_dollars_report, \
    create_cash_report, create_inventory_report, create_product_sales_report, create_invoices_report, \
    electronic_invoice_report, clients_report, electronic_invoice_report_by_invoice, filter_company
from django.db.models import Count, Sum, F, Q, Value, CharField, ExpressionWrapper, \
    DecimalField, IntegerField, When, Case, Window
from django.db.models.functions import Coalesce

from django.http import HttpResponse
from openpyxl import Workbook
from user_control.views import add_user_activity


class ReportExporter(APIView):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def post(self, request):
        start_date = request.data.get("start_date", None)
        end_date = request.data.get("end_date", None)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_ventas_{start_date}_al_{end_date}.xlsx"'

        if not start_date or not end_date:
            return Response({"error": "Debe ingresar un rango de fechas"})

        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

        wb = Workbook()
        ws = wb.active
        ws.title = "REPORTE DIARIO"
        ws.column_dimensions[get_column_letter(2)].width = 29

        invoices_queryset = filter_company(Invoice.objects.all(), self.request.user.company_id)

        terminals_report_data = (
            invoices_queryset.select_related("PaymentMethods", "payment_terminal", "created_by")
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(payment_methods__name__in=["debitCard", "creditCard"])
            .values_list("payment_terminal__name", "sale_by__fullname")
            .annotate(
                quantity=Count("id"),
                total=Sum("payment_methods__paid_amount")
            )
        )

        last_row_cards = create_terminals_report(ws, terminals_report_data, start_date, end_date)

        dollar_report_data = (
            invoices_queryset.select_related("InvoiceItems", "created_by")
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(is_dollar=True)
            .filter(invoice_items__is_gift=False)
            .values_list("sale_by__fullname")
            .annotate(
                quantity=Sum("invoice_items__usd_amount")
            )
        )

        last_row_dollars = create_dollars_report(ws, dollar_report_data, last_row_cards, start_date, end_date)

        cash_report_data = (
            invoices_queryset.select_related("InvoiceItems", "created_by")
            .filter(is_override=False)
            .filter(invoice_items__is_gift=False)
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .values_list("sale_by__fullname")
            .annotate(
                quantity=Sum("invoice_items__amount")
            )
        )

        dollar_report_data_in_pesos = (
            invoices_queryset.select_related("InvoiceItems", "created_by")
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(invoice_items__is_gift=False)
            .filter(is_dollar=True)
            .values_list("sale_by__fullname")
            .annotate(
                quantity=Sum("invoice_items__amount")
            )
        )

        cards_report_data = (
            invoices_queryset.select_related("PaymentMethods", "created_by")
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(payment_methods__name__in=["debitCard", "creditCard"])
            .values_list("sale_by__fullname")
            .annotate(
                total=Sum("payment_methods__paid_amount")
            )
        )

        transfers_report_data = (
            invoices_queryset.select_related("PaymentMethods", "created_by")
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(payment_methods__name__in=["nequi", "bankTransfer"])
            .values_list("sale_by__fullname")
            .annotate(
                total=Sum("payment_methods__paid_amount")
            )
        )

        last_row, last_column = create_cash_report(ws, last_row_dollars, last_row_cards,
                                                   cash_report_data, dollar_report_data_in_pesos, cards_report_data,
                                                   transfers_report_data, start_date, end_date
                                                   )

        apply_styles_to_cells(1, 1, last_column, last_row, ws, alignment=Alignment(horizontal="center"))

        wb.save(response)
        add_user_activity(request.user, f"Se descargó el reporte diario de ventas")
        return response


class InventoriesReportExporter(APIView):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def post(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_inventarios.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "REPORTE DE INVENTARIOS"

        inventories_report_data = (
            filter_company(Inventory.objects.all(), self.request.user.company_id).select_related("group")
            .filter(active=True)
            .annotate(
                upper_group_name=Upper("group__belongs_to__name"),
                upper_group_subname=Upper("group__name"),
                upper_name=Upper("name"),
                total_price_in_shops=F('total_in_shops') * F('buying_price'),
                total_price_in_storage=F('total_in_storage') * F('buying_price'),
                total_selling_price_in_shops=F('total_in_shops') * F('selling_price'),
                total_selling_price_in_storage=F('total_in_storage') * F('selling_price'),
                units=Value("COP", output_field=CharField())
            )
            .values_list("upper_group_name", "upper_group_subname", "code", "upper_name", "total_in_storage",
                         "total_in_shops", "buying_price", "selling_price",
                         "total_price_in_shops", "total_price_in_storage", "total_selling_price_in_shops",
                         "total_selling_price_in_storage", "units"
                         )
        )

        create_inventory_report(ws, inventories_report_data)

        wb.save(response)

        add_user_activity(request.user, f"Se descargó el reporte de inventarios")
        return response


class ItemsReportExporter(APIView):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def post(self, request):
        start_date = request.data.get("start_date", None)
        end_date = request.data.get("end_date", None)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename="reporte_ventas_x_producto_{start_date}_{end_date}.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "REPORTE DE VENTAS POR PRODUCTO"

        if not start_date or not end_date:
            return Response({"error": "Debe ingresar un rango de fechas"})

        invoices_queryset = filter_company(Invoice.objects.all(), self.request.user.company_id)
        report_data = (
            invoices_queryset.select_related("InvoiceItems")
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(invoice_items__is_gift=False)
            .values_list("invoice_items__item_code", "invoice_items__item_name")
            .annotate(
                quantity=Sum("invoice_items__quantity"),
            )
        )

        report_data_nulled = (
            invoices_queryset.select_related("InvoiceItems")
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=True)
            .filter(invoice_items__is_gift=False)
            .values_list("invoice_items__item_code", "invoice_items__item_name")
            .annotate(
                quantity=Sum("invoice_items__quantity"),
            )
        )

        report_data_gifts = (
            filter_company(Invoice.objects.all(), self.request.user.company_id).select_related("InvoiceItems")
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(invoice_items__is_gift=True)
            .values_list("invoice_items__item_code", "invoice_items__item_name")
            .annotate(
                quantity=Sum("invoice_items__quantity"),
            )
        )

        create_product_sales_report(ws, report_data, report_data_nulled, report_data_gifts, start_date, end_date,
                                    self.request.user.company)

        wb.save(response)

        add_user_activity(request.user, f"Se descargó el reporte de ventas de productos")
        return response


class InvoicesReportExporter(APIView):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def post(self, request):
        start_date = request.data.get("start_date", None)
        end_date = request.data.get("end_date", None)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_facturas_{start_date}_{end_date}.xlsx"'

        if not start_date or not end_date:
            return Response({"error": "Debe ingresar un rango de fechas"})

        wb = Workbook()
        ws = wb.active
        ws.title = "REPORTE DE FACTURACION"

        inventories_report_data = (
            filter_company(Invoice.objects.all(),
                           self.request.user.company_id).select_related("payment_terminal", "InvoiceItems", "created_by")
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(invoice_items__is_gift=False)
            .annotate(
                total_invoice=Sum("invoice_items__amount"),
            )
            .values_list(
                "created_at__date", "sale_by__fullname", "invoice_number", "dian_resolution__document_number",
                "payment_terminal__name", "total_invoice",
                "customer__document_id", "customer__name", "customer__email", "customer__phone",
                "customer__address"
            )
        )

        create_invoices_report(ws, inventories_report_data)

        wb.save(response)

        add_user_activity(request.user, f"Se descargó el reporte de facturación")
        return response


class ElectronicInvoiceExporter(APIView):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def post(self, request):
        start_date = request.data.get("start_date", None)
        end_date = request.data.get("end_date", None)

        start = datetime.strptime(start_date, "%Y-%m-%d %H:%M:%S")
        end = datetime.strptime(end_date, "%Y-%m-%d %H:%M:%S")

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file_name = 'FormatoFacturaElectronica-' + start.strftime("%Y-%m-%d_%H_%M_%S") + '-' + end.strftime(
            "%Y-%m-%d_%H_%M_%S") + '.xlsx'

        response['Content-Disposition'] = 'attachment; filename=' + file_name

        payment = {"cash": "Efectivo", "debitCard": "Tarjeta Debito Ventas", "creditCard": "Tarjeta Credito Ventas 271",
                   "nequi": "Transferencias", "bankTransfer": "Transferencias"}
        bodega = {"Guasá": "San Pablo", "CHOCOLATE": "San Pablo", "REALIDAD VIRTUAL": "San Pablo"}

        payment_conditions = [When(payment_methods__name=key, then=Value(value)) for key, value in payment.items()]
        bodega_value = [When(invoice_items__item__cost_center=key, then=Value(value)) for key, value in bodega.items()]

        if not start_date or not end_date:
            return Response({"error": "Por favor ingresar una rango de fechas correcto"})

        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"

        electronic_invoice_report_data = (
            filter_company(Invoice.objects.all(), self.request.user.company_id).select_related("invoice_number", "PaymentMethods", "payment_terminal", "InvoiceItems",
                                           "created_by")
            .filter(created_at__range=(start, end))
            .filter(is_override=False)
            .filter(invoice_items__is_gift=False)
            .annotate(
                descuento=ExpressionWrapper(
                    F("invoice_items__discount") / 100.0, output_field=DecimalField(decimal_places=2)
                ),
                valor_unitario=ExpressionWrapper(
                    F("invoice_items__item__selling_price") / 1.19, output_field=DecimalField(decimal_places=2)
                ),
                metodo_pago=Case(*payment_conditions, output_field=CharField()),
                nueva_bodega=Case(*bodega_value, output_field=CharField())
            )
            .annotate(
                null_value=ExpressionWrapper(Value(None, output_field=CharField()), output_field=CharField()),
                empresa=ExpressionWrapper(Value('SIGNOS STUDIO S.A.S.'), output_field=CharField()),
                tipo_doc=ExpressionWrapper(Value('FV'), output_field=CharField()),
                prefijo=ExpressionWrapper(Value('FESS'), output_field=CharField()),
                doc_vendedor=ExpressionWrapper(Value('832004603'), output_field=IntegerField()),
                nota=ExpressionWrapper(Value('FACTURA DE VENTA'), output_field=CharField()),
                verificado=Value(0, output_field=IntegerField()),
                medida=ExpressionWrapper(Value('Und.'), output_field=CharField()),
                iva=Value(Decimal('0.19'), output_field=DecimalField())
            )
            .values_list(
                "empresa", "tipo_doc", "prefijo", "invoice_number", "created_at__date", "doc_vendedor",
                "customer__document_id", "nota", "metodo_pago", "created_at__date",
                "null_value", "null_value", "verificado", "verificado", "null_value", "null_value", "null_value",
                "null_value", "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value", "null_value", "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value", "invoice_items__item_code", "nueva_bodega", "medida",
                "invoice_items__quantity", "iva", "valor_unitario", "descuento", "created_at__date",
                "invoice_items__item_name", "invoice_items__item__cost_center",
                "null_value", "null_value", "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value", "null_value", "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value"
            )
        )

        electronic_invoice_report(ws, electronic_invoice_report_data)

        today = timezone.now().date()

        docs = {"CC": "CC", "PA": "PASAPORTE", "NIT": "NIT", "CE": "Cédula de extranjería",
                "DIE": "Documento de identificación extranjero"}

        docs_types = [When(document_type=key, then=Value(value)) for key, value in docs.items()]

        ws2 = wb.create_sheet(title="FormatoTerceros")

        clients_report_data = (
            filter_company(Customer.objects, self.request.user.company_id)
            .annotate(total_invoices=Count('customer'))
            .filter(total_invoices__gt=0)
            .filter(~Q(customer__created_at__lt=start))
            .distinct()
            .annotate(
                doc=Case(*docs_types, output_field=CharField())
            )
            .annotate(
                b2=ExpressionWrapper(Value('0'), output_field=IntegerField()),
                null_value=ExpressionWrapper(Value(None, output_field=CharField()), output_field=CharField()),
                hoy=Value(today),
                ciudad=Coalesce('city', Value('Bogota D.C.')),
                propiedad=ExpressionWrapper(Value('Cliente;'), output_field=CharField()),
                activo=ExpressionWrapper(Value('-1'), output_field=IntegerField()),
                retencion=ExpressionWrapper(Value('Persona Natural No Responsable del IVA'), output_field=CharField()),
                clas_dian=ExpressionWrapper(Value('Normal'), output_field=CharField()),
                tipo_dir=ExpressionWrapper(Value('Casa'), output_field=CharField()),
                postal=ExpressionWrapper(Value('11111'), output_field=IntegerField()),
                direccion=Coalesce('address', Value('CR 15  01 01')),
                telefono=Coalesce('phone', Value('3333333333')),
            )
            .values_list(
                "doc", "document_id", "ciudad", "name", "null_value", "null_value", "null_value", "propiedad", "activo",
                "retencion", "hoy", "b2", "clas_dian", "null_value",
                "null_value", "null_value", "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value", "null_value", "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value", "null_value", "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value", "null_value", "tipo_dir", "ciudad", "direccion",
                "activo", "telefono", "postal", "null_value", "null_value", "null_value", "email", "null_value",
                "null_value", "null_value", "null_value", "null_value"
            )
        )

        clients_report(ws2, clients_report_data)

        new_payment_conditions = [When(payment_methods__name=key, then=Value(value)) for key, value in payment.items()]

        ws3 = wb.create_sheet(title="Detalle Facturas")

        payment_methods_report = (
            filter_company(Invoice.objects.all(), self.request.user.company_id
                           ).select_related("invoice_number", "PaymentMethods", "InvoiceItems",
                                           "created_by")
            .filter(created_at__range=(start, end))
            .filter(is_override=False)
            .filter(invoice_items__is_gift=False)
            .annotate(
                row_number=Window(
                    expression=RowNumber(),
                    partition_by=['invoice_number'],
                    order_by=F('payment_methods__id').asc()
                ),
            )
            .annotate(
                payment_method_normalized=Case(*new_payment_conditions, output_field=CharField())
            )
            .values_list(
                "invoice_number", "payment_method_normalized", "row_number"
            )
            .order_by("invoice_number")
        )

        payment_methods_report = [item for item in payment_methods_report if item[2] == 1]

        amount_report = (
            filter_company(Invoice.objects.all(), self.request.user.company_id
                           ).select_related("invoice_number", "InvoiceItems")
            .filter(created_at__range=(start, end))
            .filter(is_override=False)
            .filter(invoice_items__is_gift=False)
            .annotate(
                sum_amount=Sum("invoice_items__amount")
            )
            .values_list(
                "invoice_number", "sum_amount"
            )
            .order_by("invoice_number")
        )

        electronic_invoice_report_by_invoice(ws3, payment_methods_report, amount_report)

        wb.save(response)

        add_user_activity(request.user, f"Se descargó el reporte de facturación electrónica")
        return response