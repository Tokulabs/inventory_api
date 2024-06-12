import json
from datetime import datetime, timedelta, date
from decimal import Decimal

import boto3
from django.db import transaction
from django.db.models.functions.window import RowNumber, Rank
from django.utils import timezone
from django.db.models.functions.datetime import TruncYear, TruncDay, TruncHour, TruncMinute, TruncSecond, ExtractHour, \
    ExtractDay, ExtractMonth, ExtractWeek
from django.db.models.functions.text import Upper, Concat
from openpyxl.styles import Font, Alignment
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter

from rest_framework.viewsets import ModelViewSet
from rest_framework.views import APIView
from sqlparse.sql import Case

from app_control.models import DianResolution, Goals, PaymentTerminal, Provider, Customer, PaymentMethod
from inventory_api import settings
from inventory_api.excel_manager import apply_styles_to_cells

from .serializers import (
    GoalSerializer, Inventory, InventorySerializer, InventoryGroupSerializer, InventoryGroup,
    Invoice, InvoiceSerializer,
    InvoiceItem, DianSerializer, PaymentTerminalSerializer, ProviderSerializer, UserWithAmountSerializer,
    Customer, CustomerSerializer, InvoiceSimpleSerializer
)
from rest_framework.response import Response
from rest_framework import status
from inventory_api.custom_methods import IsAuthenticatedCustom
from inventory_api.utils import CustomPagination, get_query, create_terminals_report, create_dollars_report, \
    create_cash_report, create_inventory_report, create_product_sales_report, create_invoices_report, \
    electronic_invoice_report, clients_report, electronic_invoice_report_by_invoice
from django.db.models import Count, Sum, F, Q, Value, CharField, Func, ExpressionWrapper, Subquery, OuterRef, \
    DecimalField, IntegerField, When, Case, Window
from django.db.models.functions import Coalesce, TruncMonth, Cast, Now
from user_control.models import CustomUser
import csv
import codecs
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook


class InventoryView(ModelViewSet):
    http_method_names = ('get', 'put', 'delete', 'post')
    queryset = Inventory.objects.select_related("group", "created_by")
    serializer_class = InventorySerializer
    permission_classes = (IsAuthenticatedCustom,)
    pagination_class = CustomPagination

    def get_queryset(self):
        if self.request.method.lower() != "get":
            return self.queryset

        data = self.request.query_params.dict()
        data.pop("page", None)
        keyword = data.pop("keyword", None)

        results = self.queryset.filter(**data)

        if keyword:
            search_fields = (
                "code", "created_by__fullname", "created_by__email",
                "group__name", "name"
            )
            query = get_query(keyword, search_fields)
            return results.filter(query)

        return results

    def create(self, request, *args, **kwargs):
        request.data.update({"created_by_id": request.user.id})
        return super().create(request, *args, **kwargs)

    def update(self, request, pk=None):
        inventory = Inventory.objects.filter(pk=pk).first()
        serializer = self.serializer_class(inventory, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None):
        inventory = Inventory.objects.filter(pk=pk).first()
        inventory.delete()
        return Response({"message": "Producto eliminado satisfactoriamente"}, status=status.HTTP_200_OK)

    def toggle_active(self, request, pk=None):
        inventory = Inventory.objects.filter(pk=pk).first()
        if inventory is None:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        inventory.active = not inventory.active
        inventory.save()
        serializer = self.serializer_class(inventory)
        return Response(serializer.data)


class ProviderView(ModelViewSet):
    http_method_names = ('get', 'put', 'delete', 'post')
    queryset = Provider.objects.select_related("created_by")
    serializer_class = ProviderSerializer
    permission_classes = (IsAuthenticatedCustom,)
    pagination_class = CustomPagination

    def get_queryset(self):
        data = self.request.query_params.dict()
        keyword = data.pop("keyword", None)
        data.pop("page", None)
        results = self.queryset.filter(**data)

        if keyword:
            search_fields = (
                "nit", "created_by__fullname", "created_by__email", "name"
            )
            query = get_query(keyword, search_fields)
            results = results.filter(query)

        return results.order_by('id')

    def create(self, request, *args, **kwargs):
        request.data.update({"created_by_id": request.user.id})
        return super().create(request, *args, **kwargs)

    def update(self, request, pk):
        provider = Provider.objects.filter(pk=pk).first()
        serializer = self.serializer_class(provider, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk):
        provider = Provider.objects.filter(pk=pk).first()
        provider.delete()
        return Response({"message": "Proveedor eliminado satisfactoriamente"}, status=status.HTTP_200_OK)

    def toggle_active(self, request, pk=None):
        provider = Provider.objects.filter(pk=pk).first()
        if provider is None:
            return Response({'error': 'Proveedor no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        provider.active = not provider.active
        provider.save()
        serializer = self.serializer_class(provider)
        return Response(serializer.data)


class CustomerView(ModelViewSet):
    http_method_names = ('get', 'post', 'put', 'delete')
    queryset = Customer.objects.select_related(
        "created_by")
    serializer_class = CustomerSerializer
    permission_classes = (IsAuthenticatedCustom,)
    pagination_class = CustomPagination

    def get_queryset(self):
        if self.request.method.lower() != "get":
            return self.queryset
        data = self.request.query_params.dict()
        data.pop("page", None)
        keyword = data.pop("keyword", None)

        results = self.queryset.filter(**data).order_by('id')

        if keyword:
            search_fields = (
                "created_by__fullname", "created_by__email", "document_id", "name"
            )
            query = get_query(keyword, search_fields)
            results = results.filter(query)

        return results.order_by('id')

    def create(self, request, *args, **kwargs):
        request.data.update({"created_by_id": request.user.id})
        return super().create(request, *args, **kwargs)

    def destroy(self, request, pk):
        customer = Customer.objects.filter(pk=pk).first()
        customer.delete()
        return Response({"message": "Cliente eliminado satisfactoriamente"}, status=status.HTTP_200_OK)


class InventoryGroupView(ModelViewSet):
    http_method_names = ('get', 'put', 'delete', 'post')

    queryset = InventoryGroup.objects.select_related(
        "belongs_to", "created_by").prefetch_related("inventories")
    serializer_class = InventoryGroupSerializer
    permission_classes = (IsAuthenticatedCustom,)
    pagination_class = CustomPagination

    def get_queryset(self):
        if self.request.method.lower() != "get":
            return self.queryset
        data = self.request.query_params.dict()
        data.pop("page", None)

        keyword = data.pop("keyword", None)
        results = self.queryset.filter(**data).order_by('id')

        if keyword:
            search_fields = (
                "created_by__fullname", "created_by__email", "name"
            )
            query = get_query(keyword, search_fields)
            results = results.filter(query)

        return results.annotate(
            total_items=Count('inventories')
        )

    def create(self, request, *args, **kwargs):
        request.data.update({"created_by_id": request.user.id})
        return super().create(request, *args, **kwargs)

    def update(self, request, pk=None):
        inventory_group = InventoryGroup.objects.filter(pk=pk).first()
        serializer = self.serializer_class(inventory_group, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None):
        inventory_group = InventoryGroup.objects.filter(pk=pk).first()
        inventory_group.delete()
        return Response({"message": "Categoría eliminada satisfactoriamente"}, status=status.HTTP_200_OK)

    def toggle_active(self, request, pk=None):
        inventory_group = InventoryGroup.objects.filter(pk=pk).first()

        if inventory_group is None:
            return Response({'error': 'Categoria no encontrada'}, status=status.HTTP_404_NOT_FOUND)

        if inventory_group.belongs_to is not None:
            inventory_group_father = InventoryGroup.objects.filter(pk=inventory_group.belongs_to_id).first()
            if inventory_group_father.active == False:
                return Response({'error': 'Categoria padre no está activa'}, status=status.HTTP_400_BAD_REQUEST)

        if not inventory_group.active == False:
            for group in InventoryGroup.objects.filter(belongs_to_id=pk).all():
                group.active = False
                group.save()

        inventory_group.active = not inventory_group.active
        inventory_group.save()
        serializer = self.serializer_class(inventory_group)
        return Response(serializer.data)


class PaymentTerminalView(ModelViewSet):
    http_method_names = ('get', 'post', 'put', 'delete')
    queryset = PaymentTerminal.objects.select_related("created_by")
    serializer_class = PaymentTerminalSerializer
    permission_classes = (IsAuthenticatedCustom,)
    pagination_class = CustomPagination

    def get_queryset(self):
        if self.request.method.lower() != "get":
            return self.queryset
        data = self.request.query_params.dict()
        data.pop("page", None)

        keyword = data.pop("keyword", None)
        results = self.queryset.filter(**data)

        if keyword:
            search_fields = (
                "created_by__fullname", "created_by__email", "name", "account_code"
            )
            query = get_query(keyword, search_fields)
            results = results.filter(query)

        return results.order_by('id')

    def create(self, request, *args, **kwargs):
        request.data.update({"created_by_id": request.user.id})
        return super().create(request, *args, **kwargs)

    def update(self, request, pk=None):
        terminal = PaymentTerminal.objects.filter(pk=pk).first()
        serializer = self.serializer_class(terminal, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None):
        terminal = PaymentTerminal.objects.filter(pk=pk).first()
        terminal.delete()
        return Response({"message": "Datafono eliminado satisfactoriamente"}, status=status.HTTP_200_OK)

    def toggle_active(self, request, pk=None):
        terminal = PaymentTerminal.objects.filter(pk=pk).first()
        if terminal is None:
            return Response({'error': 'Datafono no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        terminal.active = not terminal.active
        terminal.save()
        serializer = self.serializer_class(terminal)
        return Response(serializer.data)


class InvoiceView(ModelViewSet):
    http_method_names = ('get', 'post', 'put', 'delete')
    queryset = Invoice.objects.select_related(
        "created_by", "sale_by", "payment_terminal", "dian_resolution").prefetch_related("invoice_items")
    serializer_class = InvoiceSerializer
    permission_classes = (IsAuthenticatedCustom,)
    pagination_class = CustomPagination

    def get_queryset(self):
        if self.request.method.lower() != "get":
            return self.queryset

        data = self.request.query_params.dict()
        data.pop("page", None)

        keyword = data.pop("keyword", None)
        results = self.queryset.filter(**data)

        if keyword:
            search_fields = (
                "created_by__fullname", "created_by__email", "invoice_number", "dian_resolution__document_number",
            )
            query = get_query(keyword, search_fields)
            results = results.filter(query)

        return results

    def create(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                dian_resolution = DianResolution.objects.filter(active=True).first()
                if not dian_resolution:
                    raise Exception("Necesita una Resolución de la DIAN activa para crear facturas")

                if not request.data.get("sale_by_id"):
                    request.data.update({"sale_by_id": request.user.id})

                request.data.update({"created_by_id": request.user.id})

                new_current_number = dian_resolution.current_number + 1
                dian_resolution_document_number = dian_resolution.id
                dian_resolution.current_number = new_current_number
                dian_resolution.save()

                request.data.update(
                    {"dian_resolution_id": dian_resolution_document_number, "invoice_number": new_current_number})
                super().create(request, *args, **kwargs)

                return Response({"message": "Factura creada satisfactoriamente"}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None):
        invoice = Invoice.objects.filter(pk=pk).first()
        invoice.delete()
        return Response({"message": "Factura eliminada satisfactoriamente"}, status=status.HTTP_200_OK)


class InvoiceSimpleListView(ModelViewSet):
    http_method_names = ('get',)
    permission_classes = (IsAuthenticatedCustom,)
    pagination_class = CustomPagination
    serializer_class = InvoiceSimpleSerializer
    queryset = Invoice.objects.select_related(
        "created_by", "sale_by", "payment_terminal", "dian_resolution").prefetch_related("invoice_items")

    def get_queryset(self):
        data = self.request.query_params.dict()
        data.pop("page", None)

        keyword = data.pop("keyword", None)
        results = Invoice.objects.select_related("created_by", "sale_by", "payment_terminal"
                                                 ).prefetch_related("invoice_items", "payment_methods"
                                                                    ).filter(**data).filter(
            invoice_items__is_gift=False)

        if keyword:
            search_fields = (
                "created_by__fullname", "created_by__email", "invoice_number", "dian_resolution__document_number",
            )
            query = get_query(keyword, search_fields)
            results = results.filter(query)

        return results.annotate(
            total_sum=Sum("invoice_items__amount"),
            total_sum_usd=Sum("invoice_items__usd_amount")
        ).order_by('-created_at')


class UpdateInvoiceView(APIView):
    def patch(self, request, invoice_number):
        try:
            invoice = Invoice.objects.get(invoice_number=invoice_number)
        except Invoice.DoesNotExist:
            return Response({"error": "Factura no encontrada"}, status=status.HTTP_404_NOT_FOUND)

        if invoice.is_override:
            return Response({"error": "La Factura ya está anulada"}, status=status.HTTP_400_BAD_REQUEST)

        # Actualizar el estado is_override de la factura a True
        invoice.is_override = True

        # Restaurar la cantidad de elementos en el inventario para los InvoiceItems correspondientes
        for item in invoice.invoice_items.all():
            inventory_item = item.item
            inventory_item.total_in_shops += item.quantity
            inventory_item.save()

        # Guardar los cambios en la base de datos
        invoice.save()

        return Response({"message": "Factura actualizada satisfactoriamente"}, status=status.HTTP_200_OK)


class SummaryView(ModelViewSet):
    http_method_names = ('get',)
    permission_classes = (IsAuthenticatedCustom,)
    queryset = InventoryView.queryset

    def list(self, request, *args, **kwargs):
        total_inventory = InventoryView.queryset.filter(active=True).count()
        total_group = InventoryGroupView.queryset.filter(active=True).count()
        total_users = CustomUser.objects.filter(is_superuser=False).filter(is_active=True).count()

        return Response({
            "total_inventory": total_inventory,
            "total_group": total_group,
            "total_users": total_users
        })


class InvoicePainterView(ModelViewSet):
    http_method_names = ('get',)
    permission_classes = (IsAuthenticatedCustom,)

    def list(self, request, *args, **kwargs):
        invoice_number = request.query_params.get("invoice_number", None)
        if not id:
            return Response({"error": "Debe ingresar un número de factura"}, status=status.HTTP_400_BAD_REQUEST)
        else:
            invoice = Invoice.objects.select_related(
                "payment_terminal", "created_by"
            ).filter(invoice_number=invoice_number).first()

            if not invoice:
                return Response({"error": "Factura no encontrada"}, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response(InvoiceSerializer(invoice).data)


class SalePerformance(ModelViewSet):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def top_selling(self, request, *args, **kwargs):
        query = Inventory.objects.all()
        start_date = request.data.get("start_date", None)
        end_date = request.data.get("end_date", None)

        if start_date or end_date:
            if start_date and not end_date:
                return Response({"error": "Debe ingresar una fecha de fin"}, status=status.HTTP_400_BAD_REQUEST)
            if not start_date and end_date:
                return Response({"error": "Debe ingresar una fecha de inicio"}, status=status.HTTP_400_BAD_REQUEST)
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            query = query.filter(inventory_invoices__invoice__created_at__date__gte=start_date,
                                 inventory_invoices__invoice__created_at__date__lte=end_date,
                                 inventory_invoices__invoice__is_override=False, inventory_invoices__is_gift=False)
        else:
            query = query.filter(inventory_invoices__invoice__is_override=False, inventory_invoices__is_gift=False)

        items = query.values("name", "photo").annotate(
            sum_top_ten_items=Coalesce(
                Sum("inventory_invoices__quantity"), 0
            )
        ).order_by('-sum_top_ten_items')[0:10]

        return Response(items)


class HourlySalesQuantities(ModelViewSet):
    http_method_names = ('get',)
    permission_classes = (IsAuthenticatedCustom,)

    def list(self, request, *args, **kwargs):
        hours = [{'time': hour, 'total_quantity': 0} for hour in range(24)]

        data = (
            Invoice.objects.all()
            .filter(is_override=False)
            .filter(invoice_items__is_gift=False)
            .filter(created_at__gte=datetime.now().date())
            .annotate(
                time=ExtractHour('created_at')
            ).values(
                'time'
            ).annotate(
                total_quantity=Sum('invoice_items__quantity')
            ).order_by('time')
        )

        sales_dict = {item['time']: item['total_quantity'] for item in data}

        # Update the hours list with sales data
        for hour in hours:
            if hour['time'] in sales_dict:
                hour['total_quantity'] = sales_dict[hour['time']]

        return Response(hours)


class SalesBySelectedTimeframeSummary(ModelViewSet):
    http_method_names = ('get',)
    permission_classes = (IsAuthenticatedCustom,)

    def list(self, request, *args, **kwargs):
        timeframe = request.GET.get('type', None)

        if timeframe == 'daily':
            days = []
            for i in range(7):
                date_begin = datetime.now() - timedelta(days=i)
                day = f'{date_begin.day}/{date_begin.month}'
                days.append({'day': day, 'total_amount': 0})

            days.reverse()

            data = (
                Invoice.objects.all()
                .filter(is_override=False)
                .filter(invoice_items__is_gift=False)
                .filter(created_at__gte=datetime.now().date() - timedelta(days=7))
                .annotate(
                    day=Concat(
                        ExtractDay('created_at'),
                        Value('/'),
                        ExtractMonth('created_at'),
                        output_field=CharField()
                    )
                ).values(
                    'day'
                ).annotate(
                    total_amount=Sum('invoice_items__amount')
                ).order_by('day')
            )

            sales_dict = {item['day']: item['total_amount'] for item in data}

            for day in days:
                if day['day'] in sales_dict:
                    day['total_amount'] = sales_dict[day['day']]

            return Response(days)

        elif timeframe == 'weekly':
            weeks = []
            for i in range(5):
                date_begin = datetime.now() - timedelta(weeks=i)
                week_number = f"Week {date_begin.strftime('%V')}"
                weeks.append({'week_number': week_number, 'total_amount': 0})

            weeks.reverse()

            data = (
                Invoice.objects.all()
                .filter(is_override=False)
                .filter(invoice_items__is_gift=False)
                .filter(created_at__gte=datetime.now().date() - timedelta(weeks=5))
                .annotate(
                    week_number=Concat(Value("Week "), ExtractWeek('created_at'), output_field=CharField())
                )
                .values('week_number')
                .annotate(
                    total_amount=Sum('invoice_items__amount')
                )
                .order_by('week_number')
            )

            sales_dict = {item['week_number']: item['total_amount'] for item in data}

            for week in weeks:
                if week['week_number'] in sales_dict:
                    week['total_amount'] = sales_dict[week['week_number']]

            return Response(weeks)

        elif timeframe == 'monthly':
            months = []
            month_names = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre',
                           'Octubre', 'Noviembre', 'Diciembre']
            current_year = datetime.now().year

            for i in range(1, 13):
                date_begin = datetime(current_year, i, 1)
                month_name = month_names[date_begin.month - 1]
                months.append({'month': month_name, 'total_amount': 0})

            data = (
                Invoice.objects.all()
                .filter(is_override=False)
                .filter(invoice_items__is_gift=False)
                .filter(created_at__year=current_year)
                .annotate(
                    month=ExtractMonth('created_at'),
                )
                .values('month')
                .annotate(
                    total_amount=Sum('invoice_items__amount')
                )
                .order_by('month')
            )

            sales_dict = {month_names[item['month'] - 1]: item['total_amount'] for item in data}

            for month in months:
                if month['month'] in sales_dict:
                    month['total_amount'] = sales_dict[month['month']]

            return Response(months)
        
        elif timeframe == 'general':
            today = datetime.now().date()
            current_week_start = today - timedelta(days=today.weekday())
            current_month_start = today.replace(day=1)
            current_year_start = today.replace(month=1, day=1)

            total_day = Invoice.objects.filter(
                is_override=False,
                invoice_items__is_gift=False,
                created_at__date=today
            ).aggregate(total_amount=Sum('invoice_items__amount'))['total_amount'] or 0

            total_week = Invoice.objects.filter(
                is_override=False,
                invoice_items__is_gift=False,
                created_at__gte=current_week_start
            ).aggregate(total_amount=Sum('invoice_items__amount'))['total_amount'] or 0

            total_month = Invoice.objects.filter(
                is_override=False,
                invoice_items__is_gift=False,
                created_at__gte=current_month_start
            ).aggregate(total_amount=Sum('invoice_items__amount'))['total_amount'] or 0

            total_year = Invoice.objects.filter(
                is_override=False,
                invoice_items__is_gift=False,
                created_at__gte=current_year_start
            ).aggregate(total_amount=Sum('invoice_items__amount'))['total_amount'] or 0

            general_values = {
                'diary': total_day,
                'weekly': total_week,
                'monthly': total_month,
                'annual': total_year,
            }

            return Response(general_values)
        else:
            raise Exception("Param Timeframe necesario: daily, weekly or monthly")


class SalesByUsersView(ModelViewSet):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def sales_by_user(self, request, *args, **kwargs):
        start_date = request.data.get("start_date", None)
        end_date = request.data.get("end_date", None)

        if not start_date or not end_date:
            sales_by_user = (
                Invoice.objects.select_related("InvoiceItems", "sale_by")
                .all()
                .filter(is_override=False)
                .values(
                    "sale_by__id",
                    "sale_by__fullname",
                    "sale_by__daily_goal"
                )
                .annotate(
                    total_invoice=Sum("invoice_items__amount"),
                )
            )
        else:
            sales_by_user = (
                Invoice.objects.select_related("InvoiceItems", "sale_by")
                .all()
                .filter(is_override=False)
                .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
                .values(
                    "sale_by__id",
                    "sale_by__fullname",
                    "sale_by__daily_goal"
                )
                .annotate(
                    total_invoice=Sum("invoice_items__amount"),
                )
            )

        return Response(sales_by_user)


class PurchaseView(ModelViewSet):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)
    queryset = InvoiceView.queryset

    def purchase_data(self, request, *args, **kwargs):
        query = InvoiceItem.objects.select_related("invoice", "item")
        start_date = request.data.get("start_date", None)
        end_date = request.data.get("end_date", None)

        if start_date or end_date:
            if start_date and not end_date:
                return Response({"error": "Debe ingresar una fecha de fin"}, status=status.HTTP_400_BAD_REQUEST)
            if not start_date and end_date:
                return Response({"error": "Debe ingresar una fecha de inicio"}, status=status.HTTP_400_BAD_REQUEST)
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            query = query.filter(invoice__created_at__date__gte=start_date,
                                 invoice__created_at__date__lte=end_date).filter(invoice__is_override=False)
        else:
            query = query.filter(invoice__is_override=False)

        results = query.aggregate(
            amount_total_no_gifts=Sum(F('amount'), filter=Q(is_gift=False)),
            total=Coalesce(Sum(F('quantity'), filter=Q(is_gift=False)), 0),
            gift_total=Coalesce(Sum(F('quantity'), filter=Q(is_gift=True)), 0),
            amount_total_usd=Sum(F('usd_amount'),
                                 filter=Q(invoice__is_dollar=True, invoice__is_override=False)),
            amount_total_gifts=Sum(F('amount'), filter=Q(is_gift=True))
        )

        selling_price = results.get("amount_total_no_gifts", 0)
        count = results.get("total", 0)
        gift_count = results.get("gift_total", 0)
        price_dolar = results.get("amount_total_usd", 0)
        selling_price_gifts = results.get("amount_total_gifts", 0)

        response_data = {
            "count": count,
            "gift_count": gift_count,
            "selling_price": selling_price or 0,
            "selling_price_gifts": selling_price_gifts or 0,
            "price_dolar": price_dolar or 0
        }

        return Response(response_data)


class InventoryCSVLoaderView(ModelViewSet):
    http_method_names = ('post',)
    queryset = InventoryView.queryset
    permission_classes = (IsAuthenticatedCustom,)
    serializer_class = InventorySerializer

    def create(self, request, *args, **kwargs):
        try:
            data = request.FILES['data']
        except Exception as e:
            raise Exception("Debe ingresar los productos en un archivo CSV")

        inventory_items = []

        try:
            csv_reader = csv.reader(
                codecs.iterdecode(data, 'utf-8'), delimiter=',')
            next(csv_reader)
            for row in csv_reader:
                if not row[0]:
                    continue
                inventory_items.append({
                    "group_id": int(row[0]),
                    "code": str(row[1]),
                    "name": str(row[2]),
                    "photo": str(row[3]),
                    "total_in_storage": int(row[4]),
                    "total_in_shops": int(row[5]),
                    "selling_price": float(row[6]),
                    "buying_price": float(row[7]),
                    "usd_price": float(row[8]),
                    "provider_id": int(row[9]),
                    "created_by_id": request.user.id
                })
        except csv.Error as e:
            raise Exception(e)

        if not inventory_items:
            raise Exception("El archivo CSV no puede estar vacío")

        data_validation = self.serializer_class(
            data=inventory_items, many=True)
        data_validation.is_valid(raise_exception=True)
        data_validation.save()

        return Response({
            "success": "Productos creados satisfactoriamente"
        })


class DianResolutionView(ModelViewSet):
    http_method_names = ('get', 'put', 'delete', 'post')
    queryset = DianResolution.objects.all()
    serializer_class = DianSerializer
    permission_classes = (IsAuthenticatedCustom,)
    pagination_class = CustomPagination

    def get_queryset(self):
        if self.request.method.lower() != 'get':
            return self.queryset

        current_resolution = self.queryset.filter(active=True).first()

        if current_resolution is not None and current_resolution.to_date < date.today():
            current_resolution.active = False
            current_resolution.save()

        query_set = DianResolution.objects.all()
        data = self.request.query_params.dict()
        data.pop("page", None)
        keyword = data.pop("keyword", None)

        results = query_set.filter(**data)

        if keyword:
            search_fields = (
                "created_by", "document_number", "current_number"
            )
            query = get_query(keyword, search_fields)
            return results.filter(query)

        return results

    def create(self, request, *args, **kwargs):
        request.data.update({"created_by_id": request.user.id})

        if DianResolution.objects.all().filter(active=True).exists():
            raise Exception("No puede tener más de una Resolución de la DIAN activa, "
                            "por favor, desactive primero la actual")

        return super().create(request, *args, **kwargs)

    def update(self, request, pk=None):
        dian_res = DianResolution.objects.filter(pk=pk).first()
        serializer = self.serializer_class(dian_res, data=request.data)

        if request.data.get("active") is not None and request.data.get("active", True) is True:
            if DianResolution.objects.all().filter(active=True).exists():
                raise Exception("No puede tener más de una Resolución de la DIAN activa, "
                                "por favor, desactive primero la actual")

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None):
        dian_res = DianResolution.objects.filter(pk=pk).first()
        dian_res.delete()
        return Response({"message": "Resolución DIAN eliminada satisfactoriamente"}, status=status.HTTP_200_OK)

    def toggle_active(self, request, pk=None):
        resolution = DianResolution.objects.filter(pk=pk).first()
        if resolution is None:
            return Response({'error': 'Resolución DIAN no encontrada'}, status=status.HTTP_404_NOT_FOUND)

        if DianResolution.objects.all().filter(active=True).exists() and resolution.active is False:
            raise Exception("No puede tener más de una Resolución de la DIAN activa, "
                            "por favor, desactive primero la actual")

        if resolution.to_date < date.today():
            raise Exception("No se puede activar una resolución despues de su fecha limite")

        resolution.active = not resolution.active
        resolution.save()
        serializer = self.serializer_class(resolution)
        return Response(serializer.data)


class ReportExporter(APIView):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def post(self, request):
        start_date = request.data.get("start_date", None)
        end_date = request.data.get("end_date", None)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_ventas_{start_date}_al_{end_date}.xlsx"'

        if not start_date or not end_date:
            return Response({"error": "Debe ingresar un rango de fechas"})

        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

        wb = Workbook()
        ws = wb.active
        ws.title = "REPORTE DIARIO"
        ws.column_dimensions[get_column_letter(2)].width = 29

        terminals_report_data = (
            Invoice.objects.select_related("PaymentMethods", "payment_terminal", "created_by")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(payment_methods__name__in=["debitCard", "creditCard"])
            .values_list("payment_terminal__name", "sale_by__fullname")
            .annotate(
                quantity=Count("id"),
                total=Sum("payment_methods__paid_amount")
            )
        )

        last_row_cards = create_terminals_report(ws, terminals_report_data, start_date, end_date)

        dollar_report_data = (
            Invoice.objects.select_related("InvoiceItems", "created_by")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(is_dollar=True)
            .filter(invoice_items__is_gift=False)
            .values_list("sale_by__fullname")
            .annotate(
                quantity=Sum("invoice_items__usd_amount")
            )
        )

        last_row_dollars = create_dollars_report(ws, dollar_report_data, last_row_cards, start_date, end_date)

        cash_report_data = (
            Invoice.objects.select_related("InvoiceItems", "created_by")
            .all()
            .filter(is_override=False)
            .filter(invoice_items__is_gift=False)
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .values_list("sale_by__fullname")
            .annotate(
                quantity=Sum("invoice_items__amount")
            )
        )

        dollar_report_data_in_pesos = (
            Invoice.objects.select_related("InvoiceItems", "created_by")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(invoice_items__is_gift=False)
            .filter(is_dollar=True)
            .values_list("sale_by__fullname")
            .annotate(
                quantity=Sum("invoice_items__amount")
            )
        )

        cards_report_data = (
            Invoice.objects.select_related("PaymentMethods", "created_by")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(payment_methods__name__in=["debitCard", "creditCard"])
            .values_list("sale_by__fullname")
            .annotate(
                total=Sum("payment_methods__paid_amount")
            )
        )

        transfers_report_data = (
            Invoice.objects.select_related("PaymentMethods", "created_by")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(payment_methods__name__in=["nequi", "bankTransfer"])
            .values_list("sale_by__fullname")
            .annotate(
                total=Sum("payment_methods__paid_amount")
            )
        )

        last_row, last_column = create_cash_report(ws, last_row_dollars, last_row_cards,
                                                   cash_report_data, dollar_report_data_in_pesos, cards_report_data,
                                                   transfers_report_data, start_date, end_date
                                                   )

        apply_styles_to_cells(1, 1, last_column, last_row, ws, alignment=Alignment(horizontal="center"))

        wb.save(response)
        return response


class InventoriesReportExporter(APIView):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def post(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_inventarios.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "REPORTE DE INVENTARIOS"

        inventories_report_data = (
            Inventory.objects.select_related("group")
            .filter(active=True)
            .all()
            .annotate(
                upper_group_name=Upper("group__belongs_to__name"),
                upper_group_subname=Upper("group__name"),
                upper_name=Upper("name"),
                total_price_in_shops=F('total_in_shops') * F('buying_price'),
                total_price_in_storage=F('total_in_storage') * F('buying_price'),
                total_selling_price_in_shops=F('total_in_shops') * F('selling_price'),
                total_selling_price_in_storage=F('total_in_storage') * F('selling_price'),
                units=Value("COP", output_field=CharField())
            )
            .values_list("upper_group_name", "upper_group_subname", "code", "upper_name", "total_in_storage",
                         "total_in_shops", "buying_price", "selling_price",
                         "total_price_in_shops", "total_price_in_storage", "total_selling_price_in_shops",
                         "total_selling_price_in_storage", "units"
                         )
        )

        create_inventory_report(ws, inventories_report_data)

        wb.save(response)
        return response


class ItemsReportExporter(APIView):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def post(self, request):
        start_date = request.data.get("start_date", None)
        end_date = request.data.get("end_date", None)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename="reporte_ventas_x_producto_{start_date}_{end_date}.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "REPORTE DE VENTAS POR PRODUCTO"

        if not start_date or not end_date:
            return Response({"error": "Debe ingresar un rango de fechas"})

        report_data = (
            Invoice.objects.select_related("InvoiceItems")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(invoice_items__is_gift=False)
            .values_list("invoice_items__item_code", "invoice_items__item_name")
            .annotate(
                quantity=Sum("invoice_items__quantity"),
            )
        )

        report_data_nulled = (
            Invoice.objects.select_related("InvoiceItems")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=True)
            .values_list("invoice_items__item_code", "invoice_items__item_name")
            .annotate(
                quantity=Sum("invoice_items__quantity"),
            )
        )

        report_data_gifts = (
            Invoice.objects.select_related("InvoiceItems")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(invoice_items__is_gift=True)
            .values_list("invoice_items__item_code", "invoice_items__item_name")
            .annotate(
                quantity=Sum("invoice_items__quantity"),
            )
        )

        create_product_sales_report(ws, report_data, report_data_nulled, report_data_gifts, start_date, end_date)

        wb.save(response)
        return response


class InvoicesReportExporter(APIView):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def post(self, request):
        start_date = request.data.get("start_date", None)
        end_date = request.data.get("end_date", None)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_facturas_{start_date}_{end_date}.xlsx"'

        if not start_date or not end_date:
            return Response({"error": "Debe ingresar un rango de fechas"})

        wb = Workbook()
        ws = wb.active
        ws.title = "REPORTE DE FACTURACION"

        inventories_report_data = (
            Invoice.objects.select_related("payment_terminal", "InvoiceItems", "created_by")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .annotate(
                total_invoice=Sum("invoice_items__amount"),
            )
            .values_list(
                "created_at__date", "sale_by__fullname", "invoice_number", "dian_resolution__document_number",
                "payment_terminal__name", "total_invoice",
                "customer__document_id", "customer__name", "customer__email", "customer__phone",
                "customer__address"
            )
        )

        create_invoices_report(ws, inventories_report_data)

        wb.save(response)
        return response


class ElectronicInvoiceExporter(APIView):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def post(self, request):
        start_date = request.data.get("start_date", None)
        end_date = request.data.get("end_date", None)

        start = datetime.strptime(start_date, "%Y-%m-%d %H:%M:%S")
        end = datetime.strptime(end_date, "%Y-%m-%d %H:%M:%S")

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file_name = 'FormatoFacturaElectronica-' + start.strftime("%Y-%m-%d_%H_%M_%S") + '-' + end.strftime(
            "%Y-%m-%d_%H_%M_%S") + '.xlsx'

        response['Content-Disposition'] = 'attachment; filename=' + file_name

        payment = {"cash": "Efectivo", "debitCard": "Tarjeta Debito Ventas", "creditCard": "Tarjeta Credito Ventas 271",
                   "nequi": "Transferencias", "bankTransfer": "Transferencias"}
        bodega = {"Guasá": "San Pablo", "CHOCOLATE": "San Pablo"}

        payment_conditions = [When(payment_methods__name=key, then=Value(value)) for key, value in payment.items()]
        bodega_value = [When(invoice_items__item__cost_center=key, then=Value(value)) for key, value in bodega.items()]

        if not start_date or not end_date:
            return Response({"error": "Por favor ingresar una rango de fechas correcto"})

        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"

        electronic_invoice_report_data = (
            Invoice.objects.select_related("invoice_number", "PaymentMethods", "payment_terminal", "InvoiceItems",
                                           "created_by")
            .all()
            .filter(created_at__range=(start, end))
            .filter(is_override=False)
            .filter(invoice_items__is_gift=False)
            .annotate(
                descuento=ExpressionWrapper(
                    F("invoice_items__discount") / 100.0, output_field=DecimalField(decimal_places=2)
                ),
                valor_unitario=ExpressionWrapper(
                    F("invoice_items__item__selling_price") / 1.19, output_field=DecimalField(decimal_places=2)
                ),
                metodo_pago=Case(*payment_conditions, output_field=CharField()),
                nueva_bodega=Case(*bodega_value, output_field=CharField())
            )
            .annotate(
                null_value=ExpressionWrapper(Value(None, output_field=CharField()), output_field=CharField()),
                empresa=ExpressionWrapper(Value('SIGNOS STUDIO S.A.S.'), output_field=CharField()),
                tipo_doc=ExpressionWrapper(Value('FV'), output_field=CharField()),
                prefijo=ExpressionWrapper(Value('FESS'), output_field=CharField()),
                doc_vendedor=ExpressionWrapper(Value('832004603'), output_field=IntegerField()),
                nota=ExpressionWrapper(Value('FACTURA DE VENTA'), output_field=CharField()),
                verificado=Value(0, output_field=IntegerField()),
                medida=ExpressionWrapper(Value('Und.'), output_field=CharField()),
                iva=Value(Decimal('0.19'), output_field=DecimalField())
            )
            .values_list(
                "empresa", "tipo_doc", "prefijo", "invoice_number", "created_at__date", "doc_vendedor",
                "customer__document_id", "nota", "metodo_pago", "created_at__date",
                "null_value", "null_value", "verificado", "verificado", "null_value", "null_value", "null_value",
                "null_value", "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value", "null_value", "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value", "invoice_items__item_code", "nueva_bodega", "medida",
                "invoice_items__quantity", "iva", "valor_unitario", "descuento", "created_at__date",
                "invoice_items__item_name", "invoice_items__item__cost_center",
                "null_value", "null_value", "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value", "null_value", "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value"
            )
        )

        electronic_invoice_report(ws, electronic_invoice_report_data)

        today = timezone.now().date()

        docs = {"CC": "CC", "PA": "PASAPORTE", "NIT": "NIT", "CE": "Cédula de extranjería",
                "DIE": "Documento de identificación extranjero"}

        docs_types = [When(document_type=key, then=Value(value)) for key, value in docs.items()]

        ws2 = wb.create_sheet(title="FormatoTerceros")

        clients_report_data = (
            Customer.objects.annotate(total_invoices=Count('customer'))
            .filter(total_invoices__gt=0)
            .filter(~Q(customer__created_at__lt=start))
            .distinct()
            .annotate(
                doc=Case(*docs_types, output_field=CharField())
            )
            .annotate(
                b2=ExpressionWrapper(Value('0'), output_field=IntegerField()),
                null_value=ExpressionWrapper(Value(None, output_field=CharField()), output_field=CharField()),
                hoy=Value(today),
                ciudad=Coalesce('city', Value('Bogota D.C.')),
                propiedad=ExpressionWrapper(Value('Cliente;'), output_field=CharField()),
                activo=ExpressionWrapper(Value('-1'), output_field=IntegerField()),
                retencion=ExpressionWrapper(Value('Persona Natural No Responsable del IVA'), output_field=CharField()),
                clas_dian=ExpressionWrapper(Value('Normal'), output_field=CharField()),
                tipo_dir=ExpressionWrapper(Value('Casa'), output_field=CharField()),
                postal=ExpressionWrapper(Value('11111'), output_field=IntegerField()),
                direccion=Coalesce('address', Value('CR 15  01 01')),
                telefono=Coalesce('phone', Value('3333333333')),
            )
            .values_list(
                "doc", "document_id", "ciudad", "name", "null_value", "null_value", "null_value", "propiedad", "activo",
                "retencion", "hoy", "b2", "clas_dian", "null_value",
                "null_value", "null_value", "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value", "null_value", "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value", "null_value", "null_value", "null_value", "null_value", "null_value", "null_value",
                "null_value", "null_value", "tipo_dir", "ciudad", "direccion",
                "activo", "telefono", "postal", "null_value", "null_value", "null_value", "email", "null_value",
                "null_value", "null_value", "null_value", "null_value"
            )
        )

        clients_report(ws2, clients_report_data)

        new_payment_conditions = [When(payment_methods__name=key, then=Value(value)) for key, value in payment.items()]

        ws3 = wb.create_sheet(title="Detalle Facturas")

        payment_methods_report = (
            Invoice.objects.select_related("invoice_number", "PaymentMethods", "InvoiceItems",
                                           "created_by")
            .all()
            .filter(created_at__range=(start, end))
            .filter(is_override=False)
            .filter(invoice_items__is_gift=False)
            .annotate(
                row_number=Window(
                    expression=RowNumber(),
                    partition_by=['invoice_number'],
                    order_by=F('payment_methods__id').asc()
                ),
            )
            .annotate(
                payment_method_normalized=Case(*new_payment_conditions, output_field=CharField())
            )
            .values_list(
                "invoice_number", "payment_method_normalized", "row_number"
            )
            .order_by("invoice_number")
        )

        payment_methods_report = [item for item in payment_methods_report if item[2] == 1]

        amount_report = (
            Invoice.objects.select_related("invoice_number", "InvoiceItems")
            .all()
            .filter(created_at__range=(start, end))
            .filter(is_override=False)
            .filter(invoice_items__is_gift=False)
            .annotate(
                sum_amount=Sum("invoice_items__amount")
            )
            .values_list(
                "invoice_number", "sum_amount"
            )
            .order_by("invoice_number")
        )

        electronic_invoice_report_by_invoice(ws3, payment_methods_report, amount_report)

        wb.save(response)

        return response


class GoalView(ModelViewSet):
    http_method_names = ('get', 'post', 'put', 'delete')
    queryset = Goals.objects.all()
    serializer_class = GoalSerializer
    permission_classes = (IsAuthenticatedCustom,)

    def get_queryset(self):
        if self.request.method.lower() != 'get':
            return self.queryset

        query_set = Goals.objects.all()
        data = self.request.query_params.dict()
        keyword = data.pop("keyword", None)

        results = query_set.filter(**data)

        if keyword:
            search_fields = (
                "created_by", "goal_type"
            )
            query = get_query(keyword, search_fields)
            return results.filter(query)

        return results

    def create(self, request, *args, **kwargs):
        request.data.update({"created_by_id": request.user.id})
        return super().create(request, *args, **kwargs)

    def update(self, request, pk=None):
        goal = Goals.objects.filter(pk=pk).first()
        serializer = self.serializer_class(goal, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None):
        goal = Goals.objects.filter(pk=pk).first()
        goal.delete()
        return Response({"message": "Meta eliminada satisfactoriamente"}, status=status.HTTP_200_OK)


class InvoicePaymentMethodsView(APIView):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def post(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                invoice_id = request.GET.get('invoice_id', None)

                invoice = Invoice.objects.filter(id=invoice_id).first()
                if invoice is None:
                    raise Exception("Factura no encontrada")

                invoice.payment_methods.all().delete()

                payment_methods = request.data.get("payment_methods", None)
                if not payment_methods:
                    raise Exception("Debe ingresar los métodos de pago")

                for method in payment_methods:
                    if (method.get("name") is None or method.get("paid_amount") is None or method.get(
                            "back_amount") is None or method.get("received_amount") is None):
                        raise Exception(
                            "Los métodos de pago deben tener nombre, monto pagado, monto de vuelto y monto recibido")

                for method in payment_methods:
                    PaymentMethod.objects.create(
                        invoice_id=invoice_id,
                        name=method.get("name"),
                        paid_amount=method.get("paid_amount"),
                        back_amount=method.get("back_amount"),
                        received_amount=method.get("received_amount"),
                        transaction_code=method.get("transaction_code", None)
                    )

                if 'payment_terminal_id' in request.data:
                    invoice.payment_terminal_id = request.data.get("payment_terminal_id", None)
                    invoice.save()

                if 'is_dollar' in request.data:
                    invoice.is_dollar = request.data.get("is_dollar", None)
                    invoice.save()

                return Response(
                    {"message": "Métodos de pago actualizados satisfactoriamente"},
                    status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class UploadFileView(ModelViewSet):
    http_method_names = ["post"]
    permission_classes = (IsAuthenticatedCustom, )

    def upload_photo(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No se ha proporcionado un archivo"},
                status=status.HTTP_400_BAD_REQUEST
            )

        bucket_name = settings.AWS_STORAGE_BUCKET_NAME
        aws_access_key = settings.AWS_ACCESS_KEY_ID
        aws_secret_key = settings.AWS_SECRET_ACCESS_KEY
        aws_region = settings.AWS_REGION_NAME

        s3 = boto3.client(
            "s3",
            aws_access_key_id=aws_access_key,
            aws_secret_access_key=aws_secret_key,
            region_name=aws_region
        )

        try:
            response = s3.generate_presigned_post(
                bucket_name,
                file.name,
                Fields={
                    "Content-Type": file.content_type
                },
                Conditions=[
                    {"Content-Type": file.content_type}
                ],
                ExpiresIn=30000
            )

            data = {
                "final_url": f"{response['url']}{file.name.replace(' ', '+')}",
                "endpoint_data": response
            }
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response(data)
