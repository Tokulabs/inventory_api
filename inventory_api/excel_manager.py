from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def add_values_to_col_multiple_rows(
        column,
        start_row,
        values,
        ws
):
    column = get_column_letter(column)
    for index, value in enumerate(values):
        ws[f"{column}{start_row + index}"] = value


def add_values_to_row_multiple_columns(
        start_column,
        row,
        values,
        ws
):
    for index, value in enumerate(values):
        ws[f"{get_column_letter(start_column + index)}{row}"] = value


def apply_styles_to_cells(
        start_column,
        start_row,
        end_column,
        end_row,
        ws,
        font=None,
        alignment=None,
        fill=None
):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
        for cell in row:
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if fill:
                cell.fill = fill


def sum_formula_text(
        start_row,
        end_row,
        start_column,
        end_column,
):
    return f"=SUM({get_column_letter(start_column)}{start_row}:{get_column_letter(end_column)}{end_row})"
