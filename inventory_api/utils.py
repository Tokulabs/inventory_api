import jwt
from datetime import datetime, timedelta
from django.conf import settings
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from inventory_api.excel_manager import add_values_to_row_multiple_columns, apply_styles_to_cells, sum_formula_text, \
    add_values_to_col_multiple_rows
from user_control.models import CustomUser
from rest_framework.pagination import PageNumberPagination
import re
from django.db.models import Q

# Excel styles
title_report_font: Font = Font(color="FFFFFF", bold=True)
title_report_fill = PatternFill(start_color="2262C9", end_color="2262C9", fill_type="solid")
headers_font = Font(color="000000", bold=True)
headers_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
totals_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
date_fill = PatternFill(start_color="37B350", end_color="37B350", fill_type="solid")
date_font = Font(color="FFFFFF", bold=True)
alignment = Alignment(horizontal='center', vertical='center')
border_style = Border(left=Side(border_style='thin'),
                      right=Side(border_style='thin'),
                      top=Side(border_style='thin'),
                      bottom=Side(border_style='thin'))


def get_access_token(payload, days):
    token = jwt.encode(
        {"exp": datetime.now() + timedelta(days=days), **payload},
        settings.SECRET_KEY,
        algorithm="HS256"
    )

    return token


def decodeJWT(bearer):
    if not bearer:
        return None

    token = bearer[7:]

    try:
        decoded = jwt.decode(
            token, key=settings.SECRET_KEY, algorithms="HS256"
        )

    except Exception:
        return None

    if decoded:
        try:
            return CustomUser.objects.get(id=decoded["user_id"])
        except Exception:
            return None


class CustomPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = None

    def get_page_size(self, request):
        """
        Return the page size to use for this request.
        """
        if 'page' in request.query_params:
            # If a specific page is requested, return the configured page size
            return self.page_size
        else:
            # If no page is requested, return a large value to return all results
            return 10000


def normalize_query(query_string, findterms=re.compile(r'"([^"]+)"|(\S+)').findall,
                    normspace=re.compile(r'\s{2,}').sub):
    return [normspace(' ', (t[0] or t[1]).strip()) for t in findterms(query_string)]


def get_query(query_string, search_fields):
    query = None  # Query to search for every search term
    terms = normalize_query(query_string)
    for term in terms:
        or_query = None  # Query to search for a given term in each field
        for field_name in search_fields:
            q = Q(**{"%s__icontains" % field_name: term})
            if or_query is None:
                or_query = q
            else:
                or_query = or_query | q
        if query is None:
            query = or_query
        else:
            query = query & or_query
        return query


def create_terminals_report(ws, report_data, start_date, end_date):
    if not report_data:
        return 3

    users = sorted(set([item[1].upper() for item in report_data]))
    second_row_text = ["CANT", "DATAFONO"] + list(users) + ["TOTAL DIA"]

    # add report title and date and their styles
    ws['A1'] = "VENTA EN TARJETAS"
    ws[get_column_letter(second_row_text.index("TOTAL DIA") + 1) + '1'] = f'{start_date} A {end_date}'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(users))
    apply_styles_to_cells(start_column=1, start_row=1, end_column=2 + len(users), end_row=1,
                          ws=ws, font=title_report_font, alignment=alignment, fill=title_report_fill)
    # apply style to date
    apply_styles_to_cells(start_column=second_row_text.index("TOTAL DIA") + 1, start_row=1,
                          end_column=second_row_text.index("TOTAL DIA") + 1, end_row=1,
                          ws=ws, font=date_font, alignment=alignment, fill=date_fill)

    # add values and styles to headers
    add_values_to_row_multiple_columns(1, 2, second_row_text, ws)
    apply_styles_to_cells(start_column=1, start_row=2, end_column=len(second_row_text), end_row=2, ws=ws,
                          font=headers_font, alignment=alignment, fill=headers_fill)

    # add dynamic values based in the report data
    updated_rows = []
    created_rows = 0
    terminals_drawn = {}
    for row_idx, item in enumerate(report_data, start=3):
        if item[0] not in terminals_drawn.keys():
            ws.cell(row=row_idx, column=1, value=item[2])
            ws.cell(row=row_idx, column=2, value=item[0].upper())
            ws.cell(row=row_idx, column=second_row_text.index(item[1].upper()) + 1, value=item[3])
            created_rows += 1
        else:
            ws.cell(row=terminals_drawn[item[0]], column=second_row_text.index(item[1].upper()) + 1, value=item[3])
            updated_rows.append(row_idx)
        terminals_drawn[item[0]] = row_idx

    # delete rows that were updated
    updated_rows.sort(reverse=True)
    for row_idx in updated_rows:
        ws.delete_rows(row_idx)

    new_created_rows = list(range(3, 3 + created_rows))
    # add values to total column
    total_formulas_total_column = []
    for row in new_created_rows:
        total_formulas_total_column.append(sum_formula_text(row, row, 3, second_row_text.index("TOTAL DIA")))
    add_values_to_col_multiple_rows(second_row_text.index("TOTAL DIA") + 1, new_created_rows[0],
                                    total_formulas_total_column, ws)

    # add values to total rows
    total_formulas_last_row = ["TOTAL VENTA TARJETAS"]
    for index, user in enumerate(second_row_text[2:], start=2):
        start_column = index + 1
        total_formulas_last_row.append(
            sum_formula_text(new_created_rows[0], new_created_rows[-1], start_column, start_column))

    totals_start_column = second_row_text.index("DATAFONO") + 1
    add_values_to_row_multiple_columns(totals_start_column, new_created_rows[-1] + 1, total_formulas_last_row, ws)

    # apply styles to totals row
    apply_styles_to_cells(start_column=1, start_row=new_created_rows[-1] + 1, end_column=len(second_row_text),
                          end_row=new_created_rows[-1] + 1, ws=ws,
                          font=headers_font, alignment=alignment, fill=totals_fill)

    # apply border style to the table
    for row in ws.iter_rows(min_row=1, max_row=new_created_rows[-1] + 1, min_col=1, max_col=len(second_row_text)):
        for cell in row:
            cell.border = border_style

    # apply column width to all from second_row_text[2:]
    for column in second_row_text[2:]:
        ws.column_dimensions[get_column_letter(second_row_text.index(column) + 1)].width = 23

    # apply english accounting format to all from second_row_text[2:] from new_created_rows[0] to new_created_rows[-1]
    for column in second_row_text[2:]:
        for row in new_created_rows + [new_created_rows[-1] + 1]:
            ws[f"{get_column_letter(second_row_text.index(column) + 1)}{row}"
            ].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

    return new_created_rows[-1] + 2


def create_dollars_report(ws, report_data, last_row, start_date, end_date):
    if not report_data:
        return last_row + 1

    beginning_row = last_row + 1
    users = sorted([item[0].upper() for item in report_data])
    second_row_text = ["CANT", "NOMBRE"] + users + ["TOTAL DIA"]

    # add report title and date
    ws[f'A{beginning_row}'] = "VENTAS EN DOLARES"
    ws[get_column_letter(second_row_text.index("TOTAL DIA") + 1) + f'{beginning_row}'] = f'{start_date} A {end_date}'
    ws.merge_cells(start_row=beginning_row, start_column=1, end_row=beginning_row, end_column=2 + len(users))

    # apply styles to title and date
    apply_styles_to_cells(start_column=1, start_row=beginning_row, end_column=2 + len(users), end_row=beginning_row,
                          ws=ws, font=title_report_font, alignment=alignment, fill=title_report_fill)
    apply_styles_to_cells(start_column=second_row_text.index("TOTAL DIA") + 1, start_row=beginning_row,
                          end_column=second_row_text.index("TOTAL DIA") + 1, end_row=beginning_row,
                          ws=ws, font=date_font, alignment=alignment, fill=date_fill)

    # add values and styles to headers
    add_values_to_row_multiple_columns(1, beginning_row + 1, second_row_text, ws)
    apply_styles_to_cells(start_column=1, start_row=beginning_row + 1, end_column=len(second_row_text),
                          end_row=beginning_row + 1, ws=ws, font=headers_font, alignment=alignment, fill=headers_fill)

    # add row titles to column 2
    row_titles = ["TOTAL VENTA DOLARES", "FALTANTE (MENOS)", "SOBRANTE (MAS)", "TOTAL ENTREGADO DOLARES"]
    add_values_to_col_multiple_rows(2, beginning_row + 2, row_titles, ws)

    # add dynamic values based in the report data depending on amount of users
    for item in report_data:
        ws.cell(row=beginning_row + 2, column=second_row_text.index(item[0].upper()) + 1, value=item[1])

    # add formulas to total column
    total_formulas_total_column = []
    for row in range(beginning_row + 2, beginning_row + 6):
        total_formulas_total_column.append(sum_formula_text(row, row, 3, second_row_text.index("TOTAL DIA")))
    add_values_to_col_multiple_rows(second_row_text.index("TOTAL DIA") + 1, beginning_row + 2,
                                    total_formulas_total_column, ws)

    # add formulas to total row
    total_formulas_last_row = []
    for index, user in enumerate(second_row_text[2:], start=2):
        start_column = index + 1
        total_formulas_last_row.append(
            f"={get_column_letter(start_column)}{beginning_row + 2}-{get_column_letter(start_column)}{beginning_row + 3}+{get_column_letter(start_column)}{beginning_row + 4}")
    add_values_to_row_multiple_columns(3, beginning_row + 5, total_formulas_last_row, ws)

    # apply styles to totals row
    apply_styles_to_cells(start_column=1, start_row=beginning_row + 5, end_column=len(second_row_text),
                          end_row=beginning_row + 5, ws=ws,
                          font=headers_font, alignment=alignment, fill=totals_fill)

    # apply border style to the table
    for row in ws.iter_rows(min_row=beginning_row, max_row=beginning_row + 5, min_col=1, max_col=len(second_row_text)):
        for cell in row:
            cell.border = border_style

    # apply column width to all from second_row_text[2:]
    for column in second_row_text[2:]:
        ws.column_dimensions[get_column_letter(second_row_text.index(column) + 1)].width = 23

    # apply english accounting format to all from second_row_text[2:] from beginning_row + 2 to beginning_row + 5
    for column in second_row_text[2:]:
        for row in range(beginning_row + 2, beginning_row + 6):
            ws[
                f"{get_column_letter(second_row_text.index(column) + 1)}{row}"].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

    return beginning_row + 6


def create_cash_report(ws, last_row, last_row_cards, report_data, dollar_report_data, cards_report_data,
                       transfers_report_data, start_date, end_date):
    if not report_data:
        return last_row + 1, last_row_cards + 1

    beginning_row = last_row + 1
    users = sorted([item[0].upper() for item in report_data])
    second_row_text = ["ITEM", "NOMBRE"] + users + ["TOTAL DIA"]

    # add report title and date
    ws[f'A{beginning_row}'] = "VENTAS DIARIAS"
    ws[get_column_letter(second_row_text.index("TOTAL DIA") + 1) + f'{beginning_row}'] = f'{start_date} A {end_date}'
    ws.merge_cells(start_row=beginning_row, start_column=1, end_row=beginning_row, end_column=2 + len(users))

    # apply styles to title and date
    apply_styles_to_cells(start_column=1, start_row=beginning_row, end_column=2 + len(users), end_row=beginning_row,
                          ws=ws, font=title_report_font, alignment=alignment, fill=title_report_fill)
    apply_styles_to_cells(start_column=second_row_text.index("TOTAL DIA") + 1, start_row=beginning_row,
                          end_column=second_row_text.index("TOTAL DIA") + 1, end_row=beginning_row,
                          ws=ws, font=date_font, alignment=alignment, fill=date_fill)

    # add values and styles to headers
    add_values_to_row_multiple_columns(1, beginning_row + 1, second_row_text, ws)
    apply_styles_to_cells(start_column=1, start_row=beginning_row + 1, end_column=len(second_row_text),
                          end_row=beginning_row + 1, ws=ws, font=headers_font, alignment=alignment, fill=headers_fill)

    row_titles = ["TOTAL DIA POS", "DOLAR EQUIVALENTE $ (MENOS)", "VENTAS TARJETAS (MENOS)",
                  "TRANSFERENCIAS", "OTROS DESCUENTOS (ONCES, COMIS)", "SUBTOTAL",
                  "PAGO APOYO VENTAS (MENOS)", "TOTAL EN PESOS", "FALTANTE (MENOS)", "SOBRANTE (MAS)",
                  "TOTAL ENTREGADO PESOS"]
    add_values_to_col_multiple_rows(2, beginning_row + 2, row_titles, ws)

    column_items = [
        1, 2, 3, 4, 5, "", 6, "", 7, 8
    ]
    add_values_to_col_multiple_rows(1, beginning_row + 2, column_items, ws)

    # add dynamic values based in the report data depending on amount of users
    for item in report_data:
        ws.cell(row=beginning_row + 2, column=second_row_text.index(item[0].upper()) + 1, value=item[1])

    # add dynamic values based in the dollar report data depending on amount of users
    for item in dollar_report_data:
        ws.cell(row=beginning_row + 3, column=second_row_text.index(item[0].upper()) + 1, value=item[1])

    # add dynamic values based in the cards report data depending on amount of users
    for item in cards_report_data:
        ws.cell(row=beginning_row + 4, column=second_row_text.index(item[0].upper()) + 1, value=item[1])

    # add dynamic values based in the cards report data depending on amount of users
    for item in transfers_report_data:
        ws.cell(row=beginning_row + 5, column=second_row_text.index(item[0].upper()) + 1, value=item[1])

    # add subtotal formula
    total_formulas_last_row = []
    for index, user in enumerate(second_row_text[2:], start=2):
        start_column = index + 1
        start_colum_letter = get_column_letter(start_column)
        total_formulas_last_row.append(
            f"={start_colum_letter}{beginning_row + 2}-{start_colum_letter}{beginning_row + 3}"
            f"-{start_colum_letter}{beginning_row + 4}+{start_colum_letter}{beginning_row + 5}"
            f"-{start_colum_letter}{beginning_row + 6}"
        )
    add_values_to_row_multiple_columns(3, beginning_row + 7, total_formulas_last_row, ws)

    # add values to sales support row

    # add values to total pesos row
    total_formulas_last_row = []
    for index, user in enumerate(second_row_text[2:], start=2):
        start_column = index + 1
        start_colum_letter = get_column_letter(start_column)
        total_formulas_last_row.append(
            f"={start_colum_letter}{beginning_row + 7}-{start_colum_letter}{beginning_row + 8}"
        )
    add_values_to_row_multiple_columns(3, beginning_row + 9, total_formulas_last_row, ws)

    # add values to sobrante and restante rows

    # add values to total entregado pesos row
    total_formulas_last_row = []
    for index, user in enumerate(second_row_text[2:], start=2):
        start_column = index + 1
        start_colum_letter = get_column_letter(start_column)
        total_formulas_last_row.append(
            f"={start_colum_letter}{beginning_row + 9}-{start_colum_letter}{beginning_row + 10}+"
            f"{start_colum_letter}{beginning_row + 11}"
        )
    add_values_to_row_multiple_columns(3, beginning_row + 12, total_formulas_last_row, ws)

    # add formulas to total column
    total_formulas_total_column = []
    for row in range(beginning_row + 2, beginning_row + 12):
        total_formulas_total_column.append(sum_formula_text(row, row, 3, second_row_text.index("TOTAL DIA")))
    add_values_to_col_multiple_rows(second_row_text.index("TOTAL DIA") + 1, beginning_row + 2,
                                    total_formulas_total_column, ws)

    # add styles to rows with totals
    apply_styles_to_cells(start_column=1, start_row=beginning_row + 7, end_column=len(second_row_text),
                          end_row=beginning_row + 7, ws=ws,
                          font=headers_font, alignment=alignment, fill=totals_fill)
    apply_styles_to_cells(start_column=1, start_row=beginning_row + 9, end_column=len(second_row_text),
                          end_row=beginning_row + 9, ws=ws,
                          font=headers_font, alignment=alignment, fill=totals_fill)
    apply_styles_to_cells(start_column=1, start_row=beginning_row + 12, end_column=len(second_row_text),
                          end_row=beginning_row + 12, ws=ws,
                          font=headers_font, alignment=alignment, fill=totals_fill)

    # add borders to the table
    for row in ws.iter_rows(min_row=beginning_row, max_row=beginning_row + 12, min_col=1, max_col=len(second_row_text)):
        for cell in row:
            cell.border = border_style

    # apply column width to all from second_row_text[2:]
    for column in second_row_text[2:]:
        ws.column_dimensions[get_column_letter(second_row_text.index(column) + 1)].width = 23

    # apply english accounting format to all from second_row_text[2:] from beginning_row + 2 to beginning_row + 12
    for column in second_row_text[2:]:
        for row in range(beginning_row + 2, beginning_row + 13):
            ws[
                f"{get_column_letter(second_row_text.index(column) + 1)}{row}"].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

    return beginning_row + 13, second_row_text.index("TOTAL DIA") + 1


def create_inventory_report(ws, report_data):
    row_titles = ["CATEGORIA", "GRUPO", "CODIGO", "NOMBRE", "CANTIDAD BODEGA", "CANTIDAD TIENDA", "COSTO UNIDAD",
                  "VALOR VENTA UNIDAD",
                  "COSTO TOTAL BODEGA", "COSTO TOTAL TIENDA", "TOTAL VENTA BODEGA",
                  "TOTAL VENTA TIENDA", "UNIDAD"]
    add_values_to_row_multiple_columns(1, 1, row_titles, ws)
    apply_styles_to_cells(start_column=1, start_row=1, end_column=len(row_titles), end_row=1, ws=ws,
                          font=headers_font, alignment=alignment, fill=headers_fill, border=border_style)
    for column in range(1, 13):
        ws.column_dimensions[get_column_letter(column)].width = 20

    for row_idx, row_data in enumerate(report_data, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    # apply english accounting format to all columns from 6 to 11
    for column in range(7, 13):
        for row in range(2, len(report_data) + 2):
            ws[f"{get_column_letter(column)}{row}"].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'


def create_product_sales_report(ws, report_data, report_data_nulled, report_data_gifts, start_date, end_date):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws['A1'] = "SIGNOS STUDIO SAS"

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    ws['A2'] = "NIT. 832004603-8"

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
    ws['A3'] = f"Del {start_date} al {end_date}"

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=3)

    row_titles = ["Cod.", "Descripción", "Cant."]
    add_values_to_row_multiple_columns(1, 5, row_titles, ws)
    apply_styles_to_cells(start_column=1, start_row=1, end_column=len(row_titles), end_row=5, ws=ws,
                          font=headers_font, alignment=alignment, fill=None, border=None)

    for column in range(1, 4):
        ws.column_dimensions[get_column_letter(column)].width = 15

    if report_data:
        for row_idx, row_data in enumerate(report_data, start=7):
            for col_idx, cell_value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # compute total for 3rd column from 7 to len(report_data) + 6
        total_formula = sum_formula_text(7, len(report_data) + 6, 3, 3)
        ws.cell(row=len(report_data) + 7, column=2, value="TOTAL")
        ws.cell(row=len(report_data) + 7, column=3, value=total_formula)

        apply_styles_to_cells(start_column=1, start_row=len(report_data) + 7, end_column=3, end_row=len(report_data) + 7,
                              ws=ws,
                              font=headers_font, alignment=None, fill=None, border=None)

    if report_data_nulled:
        ws.cell(row=len(report_data) + 8, column=1, value="Anulados")
        for row_idx, row_data in enumerate(report_data_nulled, start=len(report_data) + 9):
            for col_idx, cell_value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        total_formula = sum_formula_text(len(report_data) + 9, len(report_data) + 8 + len(report_data_nulled), 3, 3)
        ws.cell(row=len(report_data) + 9 + len(report_data_nulled), column=2, value="TOTAL")
        ws.cell(row=len(report_data) + 9 + len(report_data_nulled), column=3, value=total_formula)

    apply_styles_to_cells(start_column=1, start_row=len(report_data) + 8, end_column=3, end_row=len(report_data) + 8,
                          ws=ws,
                          font=headers_font, alignment=None, fill=None, border=None)
    apply_styles_to_cells(start_column=1, start_row=len(report_data) + 9 + len(report_data_nulled),
                          end_column=3, end_row=len(report_data) + 9 + len(report_data_nulled),
                          ws=ws,
                          font=headers_font, alignment=None, fill=None, border=None)

    if report_data_gifts:
        ws.cell(row=len(report_data) + 10 + len(report_data_nulled), column=1, value="Regalos")
        for row_idx, row_data in enumerate(report_data_gifts, start=len(report_data) + 11 + len(report_data_nulled)):
            for col_idx, cell_value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        total_formula = sum_formula_text(len(report_data) + 11 + len(report_data_nulled),
                                         len(report_data) + 10 + len(report_data_nulled) + len(report_data_gifts), 3, 3)
        ws.cell(row=len(report_data) + 11 + len(report_data_nulled) + len(report_data_gifts), column=2, value="TOTAL")
        ws.cell(row=len(report_data) + 11 + len(report_data_nulled) + len(report_data_gifts), column=3,
                value=total_formula)

        apply_styles_to_cells(start_column=1, start_row=len(report_data) + 10 + len(report_data_nulled),
                              end_column=3, end_row=len(report_data) + 10 + len(report_data_nulled),
                              ws=ws,
                              font=headers_font, alignment=None, fill=None, border=None)
        apply_styles_to_cells(start_column=1,
                              start_row=len(report_data) + 11 + len(report_data_nulled) + len(report_data_gifts),
                              end_column=3,
                              end_row=len(report_data) + 11 + len(report_data_nulled) + len(report_data_gifts),
                              ws=ws,
                              font=headers_font, alignment=None, fill=None, border=None)


def create_invoices_report(ws, report_data):
    row_titles = ["FECHA", "VENDEDOR", "NUMERO DE FACTURA", "DOCUMENTO DIAN", "DATAFONO", "TOTAL", "ID CLIENTE",
                  "NOMBRE CLIENTE", "EMAIL CLIENTE", "TELEFONO CLIENTE", "DIRECCION CLIENTE"]
    add_values_to_row_multiple_columns(1, 1, row_titles, ws)
    apply_styles_to_cells(start_column=1, start_row=1, end_column=len(row_titles), end_row=1, ws=ws,
                          font=headers_font, alignment=alignment, fill=headers_fill, border=border_style)
    for column in range(1, 14):
        ws.column_dimensions[get_column_letter(column)].width = 20

    for row_idx, row_data in enumerate(report_data, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    # apply english accounting format to all columns from 6 to 11
    for row in range(2, len(report_data) + 2):
        ws[f"{get_column_letter(6)}{row}"].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
